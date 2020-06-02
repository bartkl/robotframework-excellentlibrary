import openpyxl
import os
import errno
import json
from robot.api import logger
from openpyxl.utils import column_index_from_string, range_boundaries


# Base exception.
class ExcellentLibraryException(Exception):
    def __str__(self):
        return self.message


class AliasAlreadyInUseException(ExcellentLibraryException):
    def __init__(self, alias):
        self.message = ("The alias `{}' is already in use by another "
                        "workbook.".format(alias))


class ExcelFileNotFoundException(ExcellentLibraryException):
    def __init__(self, file_path):
        self.message = "file `{}' does not exist.".format(file_path)


class FileAlreadyExistsException(ExcellentLibraryException):
    def __init__(self, file_path):
        self.message = "The file `{0}' already exists.".format(file_path)


class FileAlreadyOpenedException(ExcellentLibraryException):
    def __init__(self, file_path, alias):
        self.message = ("The workbook with file_path `{0}' is already opened "
                       "with alias `{1}'.")\
                       .format(file_path, alias)


class InvalidCellCoordinatesException(ExcellentLibraryException):
    def __init__(self):
        self.message = ("Please supply sufficient coordinates for "
                        "identifying a cell.")


class NoAliasSuppliedException(ExcellentLibraryException):
    def __init__(self):
        self.message = ("Please supply an alias in order to "
                        "identify the workbook.")


class SheetExistsAlreadyException(ExcellentLibraryException):
    def __init__(self, name):
        self.message = "sheet `{}' already exists.".format(name)


class SheetNotFoundException(ExcellentLibraryException):
    def __init__(self, name):
        self.message = "Could not find sheet `{0}'.".format(name)


class TooFewColumnNamesSuppliedException(ExcellentLibraryException):
    def __init__(self,):
        self.message = ("The amount of column names supplied is smaller than "
                        "the amount of columns.")


class UnknownWorkbookException(ExcellentLibraryException):
    def __init__(self, alias):
        self.message = "No opened workbook found with alias `{}'."\
                       .format(alias)


class UnopenedWorkbookException(ExcellentLibraryException):
    def __init__(self, alias):
        self.message = \
            ("workbook with alias `{}' is unknown, please open it before "
             "using it.".format(alias))


class ExcellentLibrary:
    """This library is built on top of _OpenPyXL_ in order to bring its
    functionality to _Robot Framework_. The major motivation for this was
    to add support for _Excel 2010_ (XSLX) files, which _ExcelLibrary_ does
    not support.

    = Usage =
    *TODO*: Showcase one full-feature, gigantic test suite covering pretty much
    all functionality and variety one could possibly encounter.

    The OpenPyXL documentation is quite immature, so if you really need to
    understand the implementation better you are forced to experiment or
    read the source code.
    """


    __version__ = '0.9.3'
    ROBOT_LIBRARY_SCOPE = 'GLOBAL'

    def __init__(self):
        self.workbooks = {}
        self.active_workbook_alias = None
        self.active_workbook = None

    def _add_to_workbooks(self, file_path, workbook, alias=None):
        absolute_file_path = os.path.abspath(file_path)
        """Adds the specified workbook to the opened workbooks dictionary. The
        supplied alias will be used as the key for the dictionary entry. This 
        defaults to the file path in case no alias is given.

        The values in this dictionary are dictionaries themselves, containing
        the file path and the OpenPyXL Workbook object.

        An example of what the opened workbooks dictionary could look like:
        {   "first file": {
                "file_path": "H:\\Data\\Workbook 1.xlsx",
                "workbook":  workbook_1_object
            },
            "second file": {
                "file_path": "H:\\Data\\Workbook 2.xlsx",
                "workbook":  workbook_2_object
            }
        }
        """
        if not alias:
            alias = absolute_file_path  # Setting the default.

        if alias in self.workbooks.keys():
            raise AliasAlreadyInUseException(alias)

        for workbook_entry in self.workbooks.values()   :
            if absolute_file_path == workbook_entry["file_path"]:
                existing_alias = self._get_alias_of_workbook_by_file_path(absolute_file_path)
                raise FileAlreadyOpenedException(absolute_file_path, existing_alias)

        self.workbooks[alias] = {"file_path": absolute_file_path,
                                 "workbook": workbook}
        self._set_new_active_workbook(alias)

    def _get_alias_of_workbook_by_file_path(self, file_path):
        """Gets the alias of the workbook identified by the supplied
        file path.
        """
        for alias, workbook_entry in self.workbooks.items():
            if workbook_entry["file_path"] == file_path:
                return alias

    def _get_column_names_from_header_row(self, sheet):
        """Gets the values from the header row and returns them as a list.
        """
        column_names = []
        header_row = sheet[1]
        for cell in header_row:
            column_names.append(cell.value)
        return column_names

    def _remove_from_workbooks(self, alias):
        """Removes the workbook identified by the supplied alias.
        """
        try:
            del self.workbooks[alias]
        except AttributeError:
            pass

    def _resolve_cell_coordinates(self, locator):
        """Resolves the cell coordinates based on several possible forms
        of `locator` parameter.

        Currently supports the following locator forms:
        - __Coordinates__
        - __A1 Notation__

        For more detail see the keyword documentation of the `Read From Cell`
        keyword.
        """

        locator = locator.strip()
        if "," in locator:  # Coordinates.
            if locator.startswith("coords:"):
                locator = locator[7:]
            locator = locator.lstrip('(').rstrip(')')
            coords_parts = locator\
                           .lstrip('(')\
                           .rstrip(')')\
                           .split(',')
            row_nr, col_nr = coords_parts[1].strip(),\
                             coords_parts[0].strip()

        else:  # Assume A1 notation.
            if locator.startswith("a1:"):
                locator = locator[3:]

            a1_col = ""
            a1_row = ""
            for char in locator:
                if char.isdigit():
                    a1_row += char
                else:
                    a1_col += char

            col_nr = column_index_from_string(a1_col)
            row_nr = a1_row

        return int(row_nr), int(col_nr)

    def _set_new_active_workbook(self, alias):
        """Sets the workbook identified by the supplied alias as the active one.
        """
        if not alias in self.workbooks.keys():
            raise UnknownWorkbookException(alias)
        self.active_workbook_alias = alias
        self.active_workbook = self.workbooks[alias]["workbook"]

    def close_all_workbooks(self):
        """Closes all opened workbooks.

        Changes made to the file won't be saved automatically.
        Use the `Save` keyword to save the changes to the file.
        """
        for alias in self.workbooks.keys():
            self.close_workbook(alias)

    def close_workbook(self, alias=None):
        """Closes the workbook identified by the supplied alias.
        
        If no alias is provided, the alias of the active workbook
        is used. In this case a new workbook becomes active.

        Changes made to the file won't be saved automatically.
        Use the `Save` keyword to save the changes to the file.
        """
        if not alias:
            alias = self.active_workbook_alias
        try:
            if alias == self.active_workbook_alias:
                set_new_active_workbook = True
            else:
                set_new_active_workbook = False

            self.workbooks[alias]["workbook"].close()
            self._remove_from_workbooks(alias)

            if set_new_active_workbook and len(self.workbooks) > 0:
                print(self.workbooks.keys())
                new_alias = list(self.workbooks.keys())[0]
                self._set_new_active_workbook(new_alias)

        except KeyError:
            logger.warning("Cannot close workbook with alias `{}': workbook "
                            "not opened.".format(alias))

    def create_sheet(self, sheet_name):
        """Creates a sheet in the active workbook.

        The ``name`` parameter must be used to supply the name of the sheet.
        If the sheet already exists, a ``SheetAlreadyExistsException`` is
        raised.
        """
        if not sheet_name in self.active_workbook.sheetnames:
            self.active_workbook.create_sheet(title=sheet_name)
        else:
            raise SheetExistsAlreadyException(sheet_name)

    def create_workbook(self, file_path, overwrite_file_if_exists=False, alias=None):
        """Creates a new workbook and saves it to the given file path.

        The file will also be considered opened, i.e. it will be added to the
        internal dictionary of opened workbooks using the supplied alias. 
        If no alias is supplied, it will default to the file path.

        In case the given file already exists, an ``FileAlreadyExistsException`` is raised.
        If you wish to overwrite the existing file, pass the argument ``overwrite_file_if_exists=${TRUE}``.


        _NOTE_: It is advised to supply an absolute path to avoid confusion.

        Examples:

        |  Create workbook  | H:\\Workbook 1.xlsx  |  # `alias` defaults to absolute file path  |
        |  Create workbook  | H:\\Workbook 2.xlsx  |  alias=second workbook                     |

        """
        workbook = openpyxl.Workbook()
        if os.path.isfile(file_path) and not overwrite_file_if_exists:
            raise FileAlreadyExistsException(file_path)
        else:
            workbook.save(file_path)

        # Add it to the opened workbooks dictionary.
        self._add_to_workbooks(file_path, workbook, alias=alias)

    def get_column_count(self):
        """Returns the number of non-empty columns in the active sheet.
        """
        sheet = self.active_workbook.active
        return sheet.max_column

    def get_row_count(self):
        """Returns the number of non-empty rows in the active sheet.
        """
        sheet = self.active_workbook.active
        return sheet.max_row

    def get_row_iterator(self):
        """Returns an iterator for looping over the rows in the active sheet.

        _NOTE_: This won't be needed often and it is advised to avoid this as much as
        possible, since it is unfriendly to read and hacky in its use with
        respect to Robot Framework.
        """
        sheet = self.active_workbook.active
        return sheet.iter_rows()

    def log_opened_workbooks(self, to_log=True, to_console=False):
        """Logs the dictionary in which the opened workbooks are kept.

        If ``to_log`` is ``True``, this keyword outputs in the log file.

        If ``to_console`` is ``True``, this keyword outputs on the console.

        Note that it is perfectly fine to log to both the log file and console simultaneously.
        """
        if to_log:
            logger.info(self.workbooks)
        if to_console:
            logger.console(self.workbooks)

    def open_workbook(self, file_path, alias=None, keep_vba=False):
        """Opens the workbook found at the given file path.
        _NOTE_: Please note that at present _XLS_ files are not supported.

        The file will be added to the internal dictionary of opened workbooks
        using the supplied alias. If no alias is supplied, it will default to
        the file path. The opened workbook will also be made the active workbook.

        The ``file_path`` parameter should point to the location of the file on
        the filesystem. It is advisable to make this an absolute path to avoid
        confusion.

        The ``alias`` can be used to give a more practical name to your workbook,
        which comes in handy when working with several opened workbooks simultaneously.

        If the file you want to open contains VBA (macros), please pass ``keep_vba=${TRUE}``
        in order to preserve the VBA code.

        *Warning*: make sure to explicitly switch to the sheet you want to
        work with by using the `Switch sheet` keyword. Contrary to expectations,
        the active sheet by default is not necessarily the first one in tab-order.

        Examples:
        |  Open workbook                            |  H:\\Data\\Wb1.xlsx           |                       |                   |  # now _Wb1.xlsx_ is the active workbook  |
        |  Open workbook                            |  H:\\Data\\Wb2.xlsx           |  alias=wb2            |                   |  # Now _Wb2.xlsx_ is the active workbook. |
        |  Switch workbook                          |  H:\\Data\\Wb1.xlsx           |                       |                   |  # now _Wb1.xlsx_ is the active workbook  |
        |  Close workbook                           |  wb2                          |                       |                   |  # now _wb2_ is closed and _Wb1.xlsx_ is set to be the active workbook |
        |  Close workbook                           |                               |                       |                   |  # now _Wb1.xlsx_ is closed because it was the active workbook |
        |  Open workbook                            |  H:\\Data\\WbWithMacro.xlsx   |  alias=Macro Workbook |  keep_vba=${TRUE} |  # Macro's are preserved and properly saved on `Save`  | 

        """
        try:
            workbook = openpyxl.load_workbook(file_path, keep_vba=keep_vba)
        except IOError as e:
            if e.errno == errno.ENOENT:
                raise ExcelFileNotFoundException(file_path)
            else:
                raise e
        self._add_to_workbooks(file_path, workbook, alias=alias)

    def read_from_cell(self,
                       cell,
                       cell_obj=None,
                       trim=False):
        """Reads the data from the cell identified by the given locator.

        A cell can be identified in two ways:

        - _Coordinates_: provide both the row and column numbers of the cell, starting with 1.
        - _A1 Notation_: provide the commonly used A1 Notation from Excel.
        See the examples below for more detailed use:

        |  # Coordinates. | |                           |  # no parentheses and space after comma is ok  |
        |  ${value}=  |  Read from cell  |  1, 2        |  # coords prefix is ok  |
        |  ${value}=  |  Read from cell  |  coords:1,3  |  # parentheses are fine  |
        |  ${value}=  |  Read from cell  |  (1,4)       |    |
        |  ${value}=  |  Read from cell  |  (1, 5)      |    |
        | | |                                           |    |
        |  # A1 Notation. | |                           |    |
        |  ${value}=  |  Read from cell  |  B2          |  # no prefix for a1 notations is also ok  |
        |  ${value}=  |  Read from cell  |  a1:CC2      |  # with prefix is fine as well  |

        Note that the prefixes ``coords`` and ``a1`` are optional. Without a prefix the library is still capable of resolving which locator form you intended to use. Arguably though, using them is more explicit and therefore improves readability.

        The ``cell_obj`` argument can be used to pass an _OpenPyXL
        Cell_ object to read from. This is not intended for typical use.

        By default the value read from the cell is obtained untouched, verbatim. To trim the surrounding whitespace you can pass the argument ``trim=${TRUE}``.

        """
        sheet = self.active_workbook.active

        if cell_obj:
            cell = cell_obj
        else:
            row_nr, col_nr = self._resolve_cell_coordinates(cell)
            cell = sheet.cell(row_nr, col_nr)

        if trim and cell.value is not None:
            return cell.value.strip()
        else:
            return cell.value

    def read_sheet_data(self,
                        column_names=None,
                        get_column_names_from_header_row=False,
                        cell_range=None,
                        trim=False):
        """Reads all the data from the active sheet.

        This keyword can output the sheet data in two formats:
            - _As a list of dictionaries_. In the case column names are
            supplied or obtained (see relevant parameters described below),
            the rows will be represented through dictionaries, of which the
            keys will correspond to the column names.
            - _As a list of lists_. If no column names are provided or
            obtained, each row will be read from the sheet as a list, and
            the returned data will, therefore, be a list of all such lists.

        To use column names the following two parameters can be used:

        - If ``column_names`` is provided it is expected to be a list which will
        be used to name the columns in the supplied order.

        - If ``get_column_names_from_header_row`` is ``True``, the column names
        will be read from the first row in the sheet. In this case, the first row
        will not be read as part of the sheet data.

        _NOTE_: If both parameters are supplied, the `column_names`` list
        will have precedence. You will get a warning in your log when this
        situation occurs though.

        Use ``cell_range`` if you want to get data from only that range in the
        sheet, rather than all of the data in it. The expected input form is in
        _A1 Notation_. For example: ``A1:B3``.

        If ``trim`` is ``True``, all cell values are trimmed, i.e. the
        surrounding whitespace is removed.

        Examples:

        |  Read entire sheet with column names from header row  |  |                                 |                 |
        |  Open workbook   |  ${PROPER EXCEL FILE}  |  # no alias provided: defaulting to file path  |                 |
        |  Switch sheet    |  Sheet 1 (with header) |                                                |                 |
        |  @{data sheet}=  |  Read sheet data       |  get_column_names_from_header_row=${TRUE}      |                 |
        |  :FOR            |  ${row}                |  IN                                            |  @{data sheet}  |
        |  \               |  Log list              |  ${row}                                        |                 |
        |  Close workbook  |  ${PROPER EXCEL FILE}  |                                                |                 |

        |  Read sheet range without column names (trimmed) | |                |                 |
        |  Open workbook   |  ${PROPER EXCEL FILE}  |  first excel file       |                 |
        |  Switch sheet    |  Sheet 1 (with header) |                         |                 |
        |  @{data sheet}=  |  Read sheet data       |  cell_range=A1:B3       |  trim=${TRUE}   |
        |  :FOR            |  ${row}                |  IN                     |  @{data sheet}  |
        |  \               |  Log dictionary        |  ${row}                 |                 |
        |  Close workbook  |                        |                         |                 |
    
        For more examples check out the included test suite.
        """
        sheet = self.active_workbook.active
        skip_first_row = False

        if get_column_names_from_header_row:
            if column_names:
                logger.warning("Both the `column_names' and "
                                "`get_column_names_from_header_row' "
                                "parameters were supplied. Using "
                                "`column_names' and ignoring the other.")
            else:
                skip_first_row = True
                column_names = self._get_column_names_from_header_row(sheet)

        if cell_range:
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)
            row_iterator = sheet.iter_rows(min_col=min_col, min_row=min_row,
                                           max_col=max_col, max_row=max_row)
            if column_names:
                column_names = column_names[min_col-1:max_col]
        else:
            row_iterator = sheet.iter_rows()

        if skip_first_row:
            next(row_iterator)  # Skip first row in the case of a header.

        if column_names:
            sheet_data = []
            for row in row_iterator:
                row_data = {}
                for i, cell in enumerate(row):
                    try:
                        row_data[column_names[i]] =\
                        self.read_from_cell(None, cell_obj=cell, trim=trim)
                    except IndexError:
                        raise TooFewColumnNamesSuppliedException

                if not all(value is None for value in row_data.values()):
                    sheet_data.append(row_data)
        else:
            sheet_data = []
            for row in row_iterator:
                row_data = [
                    self.read_from_cell(None, cell_obj=cell, trim=trim)
                    for cell in row
                ]
                if not all(value is None for value in row_data):
                    sheet_data.append(row_data)

        return sheet_data

    def remove_sheet(self, sheet_name):
        """Removes the sheet identified by its name from the active workbook.

        The ``name`` parameter must be used to supply the name of the sheet.
        If the sheet does not exist, a ``SheetNotFoundException`` will be
        raised.
        """
        try: 
            del self.active_workbook[sheet_name]
        except KeyError:
            raise SheetNotFoundException(sheet_name)

    def save(self, reopen_after_save=False):
        """Saves the changes to the currently active workbook.

        _NOTE_: When manipulating sheets/cells, you are working with
        object representations in memory, not the factual data on disk.
        Only when you choose to make the changes persistent by calling this
        keyword, will those changes be saved to the file.

        If the `reopen_after_save` parameter is `True` the workbook will be
        closed and then opened again after being saved. This parameter's sole
        purpose is to serve as a workaround for the corruption of
        macro-enabled workbooks with comments when saving more than once.
        (see: OpenPyXL Bitbucket issue 861)
        """
        file_path = self.workbooks[self.active_workbook_alias]["file_path"]
        self.active_workbook.save(file_path)
        if reopen_after_save:
            self.close_workbook()
            self.open_workbook(file_path, keep_vba=True)  # The workaround is
                                                          # meant to solve the
                                                          # bug with macro
                                                          # enabled workbooks,
                                                          # so `keep_vba=True`
                                                          # is fine.

    def switch_sheet(self, sheet_name):
        """Switches to the sheet with the supplied name within the active
        workbook.

        Please supply the ``sheet_name`` parameter to identify which sheet you
        want to switch to.
        """
        sheet = self.active_workbook[sheet_name]
        index = self.active_workbook.index(sheet)
        self.active_workbook.active = index

    def switch_workbook(self, alias):
        """Switches to the workbook identified by ``alias``, i.e. make
        that the active workbook.

        _NOTE_: You can only switch to workbooks which are opened. This
        keyword won't do that for you, so make sure you've opened the
        workbook you want to switch to using `Open workbook`.

        Examples:

        |  Opening several workbooks and switching between them  |
        |  Open workbook   |  ${PROPER EXCEL FILE}  |  alias=first excel file  |  # supply alias with or without `alias=`: both is fine |
        |  Open workbook   |  ${WEIRD EXCEL FILE}   |  second excel file       |  # when opening a workbook it is made the active one   |
        |  Switch workbook |  first excel file      |                          |                                                        |
        |  Switch sheet    |  Sheet 2 (no header)   |                          |                                                        |
        |  Close workbook  alias=first excel file   |                          |                                                        |
        |  Close workbook  alias=second excel file  |                          |                                                        |
        """
        try:
            self._set_new_active_workbook(alias)
        except KeyError:
            raise UnopenedWorkbookException(alias)

    def write_to_cell(self,
                      cell,
                      value,
                      number_format=None):
        """Writes a value to the supplied cell.

        For an explanation of how to identify a cell, please see the `Read From Cell` keyword documentation.
        For the sake of convenience I will stick with A1 Notation.

        Writing a value to a cell, then, is really straight-forward:

        | Write To Cell | B1   | Hello       | # this is ok! |

        It is possible to format the cell using the ``number_format`` parameter.
        In order for this to work properly with the data you're writing, you must
        make sure that the data type of the latter is compatible with what the
        number formatting expects. For example, to format a cell as a number
        that's rounded to two decimals, one should write data of a type number. To
        format a cell to hold a date-time value, a Python date-time object should 
        be passed in for it to function.

        Some examples:
        | Write To Cell | B1 | Hello      |                           | # OK  |
        | Write To Cell | B1 | ${2}       |                           | # OK  |
        | Write to cell | A1 | 1.233      | number_format=#.#         | # Bad |
        | Write to cell | A1 | ${1.233}   | number_format=#.#         | # OK  |
        | Write to cell | C1 | 2018-04-01 | number_format=yyyy-dd-mm  | # Bad |
        | ${now}        | DateTime.Get current date |       |         |       |
        | Write to cell | D1  | ${now}    |  number_format=yyyy-dd-mm | # OK  |
        | Write to cell | D1$ | {now}     |  number_format=jjjj-dd-mm | # Bad |

        _NOTE_: The ``numer_format`` parameter seems to assume the US locale, so
        make sure to delimit numbers with dots ("."), and format your dates using
        ``yyyy`` for example rather than ``jjjj`` (in Dutch). Excel will honour your
        own locale settings anyways, so don't worry about it.
        """
        sheet = self.active_workbook.active
        row_nr, col_nr = self._resolve_cell_coordinates(cell)

        cell = sheet.cell(row_nr, col_nr)

        if number_format:
            cell.number_format = number_format

        cell.value = value


